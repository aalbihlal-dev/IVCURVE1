"""
generate_report.py
Generates IV Care Test Report Excel matching the exact format of IV_Test_Full_Report.xlsx
Called from the PWA backend or run standalone.
Usage: python3 generate_report.py modules.json output.xlsx
"""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

# ── EXACT COLORS FROM TEMPLATE ────────────────────────────
DARK_BLUE   = "FF1F4E79"   # header row 1, module sheet title
MID_BLUE    = "FF2E75B6"   # header row 2, section headers
GREEN_HDR   = "FF375623"   # T1 group header
GREEN_SUB   = "FF4E6B3A"   # T1 sub-header
BROWN_HDR   = "FF7B3F00"   # T2 group header
BROWN_SUB   = "FF8B4513"   # T2 sub-header
PURPLE_HDR  = "FF4A235A"   # Delta group header
PURPLE_SUB  = "FF6C3483"   # Delta sub-header
PARAM_BLUE  = "FFBDD7EE"   # Parameter label cells (alternating A)
DATA_LIGHT  = "FFDEEAF1"   # Data cells alternating (light)
DELTA_BG    = "FFF5E6FA"   # Delta column background
ASSESS_BG   = "FFEBF3FB"   # Assessment row background
ROW_GREEN   = "FFE2EFDA"   # Summary normal row
WHITE       = "FFFFFFFF"
BLACK       = "FF000000"
RED_FONT    = "FFC00000"
DARK_GREEN_FONT = "FF006400"
BLUE_FONT   = "FF1F4E79"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=9, color=BLACK, name="Arial"):
    return Font(bold=bold, size=size, color=color, name=name)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="FFB0C4DE")
    return Border(left=s, right=s, top=s, bottom=s)

def apply(cell, value, fill_color=None, bold=False, size=9, color=BLACK,
          align="center", border=True):
    cell.value = value
    if fill_color:
        cell.fill = fill(fill_color)
    cell.font = font(bold=bold, size=size, color=color)
    cell.alignment = center() if align == "center" else left()
    if border:
        cell.border = thin_border()


def build_summary(wb, modules, test_date):
    ws = wb.create_sheet("Summary", 0)

    # ── Row 1: Title ──
    ws.merge_cells("A1:U1")
    apply(ws["A1"],
          "IV CURVE TEST REPORT — SOLAR MODULE PERFORMANCE SUMMARY",
          fill_color=DARK_BLUE, bold=True, size=14, color=WHITE)
    ws.row_dimensions[1].height = 31.5

    # ── Row 2: Subtitle ──
    ws.merge_cells("A2:U2")
    apply(ws["A2"],
          f"Test Date: {test_date}  |  Location: GMT+3  |  Device: Fluke Solmetric PV Analyzer 5.1  |  Total Modules: {len(modules)}",
          fill_color=MID_BLUE, size=10, color=WHITE)
    ws.row_dimensions[2].height = 19.5

    # ── Row 3: Headers ──
    headers = [
        "No.", "Status", "Serial Number", "Model", "MVPS", "Inv", "DCB", "String",
        "Rated Pmax (W)",
        "Without Cover Perf (%)", "Without Cover\nPmax Meas (W)", "Without Cover\nPmax Pred (W)", "Without Cover\nPmax STC (W)",
        "Backside Covered Perf (%)", "Backside Covered\nPmax Meas (W)", "Backside Covered\nPmax Pred (W)", "Backside Covered\nPmax STC (W)",
        "Bifacial Gain STC (W)", "Bifacial Gain (%)", "Alerts", "Comments"
    ]
    for c, h in enumerate(headers, 1):
        apply(ws.cell(row=3, column=c), h,
              fill_color=DARK_BLUE, bold=True, size=9, color=WHITE)
    ws.row_dimensions[3].height = 36

    # ── Column widths ──
    widths = [5,10,28,22,10,6,8,7,10,12,14,14,12,10,14,14,12,14,12,30,60]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── Data rows ──
    for i, m in enumerate(modules):
        row = i + 4
        t1 = m.get("t1") or {}
        t2 = m.get("t2") or {}
        np_ = m.get("nameplate") or {}
        mtype = m.get("type", "normal")

        dW = None
        dPct = None
        try:
            v1stc = float(t1.get("pmaxSTC") or 0)
            v2stc = float(t2.get("pmaxSTC") or 0)
        except (TypeError, ValueError):
            v1stc = v2stc = 0
        if v1stc and v2stc:
            dW   = round(v1stc - v2stc, 2)
            dPct = round(dW / v1stc * 100, 1)

        all_alerts = " | ".join(filter(None, [t1.get("alerts",""), t2.get("alerts","")])) or "None"

        status_map = {"damaged": "🔴 DAMAGED", "spare": "Spare", "normal": "Normal"}
        status = status_map.get(mtype, "Normal")

        # Row color
        if mtype == "damaged":
            row_fill = "FFFCE8E6"
        elif mtype == "spare":
            row_fill = "FFE8F0FE"
        elif all_alerts != "None":
            row_fill = "FFFFF3CD"
        else:
            row_fill = ROW_GREEN

        vals = [
            i+1, status, m.get("serialNumber",""), m.get("model",""),
            m.get("mvps",""), m.get("inverter",""), m.get("dcb",""), m.get("string",""),
            np_.get("pmax",""),
            t1.get("performance",""), t1.get("pmaxMeasured",""), t1.get("pmaxPredicted",""), t1.get("pmaxSTC",""),
            t2.get("performance","N/A"), t2.get("pmaxMeasured","N/A"), t2.get("pmaxPredicted","N/A"), t2.get("pmaxSTC","N/A"),
            dW if dW is not None else "N/A",
            dPct if dPct is not None else "N/A",
            all_alerts,
            m.get("assessment","")
        ]

        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c)
            bold_cols = {2}  # Status column bold
            apply(cell, v, fill_color=row_fill, bold=(c in bold_cols), size=8, color=BLACK)

        ws.row_dimensions[row].height = 60

    ws.freeze_panes = "A4"


def build_module_sheet(wb, m, idx):
    sheet_name = f"M{str(idx).zfill(2)}"
    ws = wb.create_sheet(sheet_name)

    t1 = m.get("t1") or {}
    t2 = m.get("t2") or {}
    np_ = m.get("nameplate") or {}
    mtype = m.get("type", "normal")
    status_map = {"damaged": "DAMAGED", "spare": "Spare", "normal": "Normal"}
    status = status_map.get(mtype, "Normal")
    sn = m.get("serialNumber", "SN N/A")

    # ── Row 1: Module title ──
    ws.merge_cells("A1:R1")
    apply(ws["A1"], f"Module {idx} — {status}  |  {sn}",
          fill_color=DARK_BLUE, bold=True, size=12, color=WHITE)
    ws.row_dimensions[1].height = 27.75

    # ── Row 2: Module Information section ──
    ws.merge_cells("A2:K2")
    apply(ws["A2"], "Module Information",
          fill_color=MID_BLUE, bold=True, size=10, color=WHITE)
    ws.row_dimensions[2].height = 15

    # ── Row 3: Module info headers ──
    info_headers = ["Serial Number","Model","MVPS","DCB","String","Inverter",
                    "Rated Pmax","Voc (nom)","Vmp (nom)","Isc (nom)","Imp (nom)"]
    for c, h in enumerate(info_headers, 1):
        apply(ws.cell(3, c), h, fill_color=MID_BLUE, bold=True, size=9, color=WHITE)
    ws.row_dimensions[3].height = 21.75

    # ── Row 4: Module info values ──
    info_vals = [
        sn, m.get("model",""), m.get("mvps",""), m.get("dcb",""),
        m.get("string",""), m.get("inverter",""),
        f"{np_.get('pmax','')}W" if np_.get('pmax') else "",
        f"{np_.get('voc','')}V" if np_.get('voc') else "",
        f"{np_.get('vmp','')}V" if np_.get('vmp') else "",
        f"{np_.get('isc','')}A" if np_.get('isc') else "",
        f"{np_.get('imp','')}A" if np_.get('imp') else "",
    ]
    for c, v in enumerate(info_vals, 1):
        apply(ws.cell(4, c), v, fill_color=WHITE, size=9, color=BLACK)
    ws.row_dimensions[4].height = 29.25

    # ── Row 5: IV Test Results section ──
    ws.merge_cells("A5:R5")
    apply(ws["A5"], "IV Test Results",
          fill_color=MID_BLUE, bold=True, size=10, color=WHITE)
    ws.row_dimensions[5].height = 18

    # ── Row 6: Group headers ──
    ws.merge_cells("A6:A7")
    apply(ws["A6"], "Parameter", fill_color=DARK_BLUE, bold=True, size=9, color=WHITE)

    ws.merge_cells("B6:E6")
    apply(ws["B6"], "Without Cover (Test 1)", fill_color=GREEN_HDR, bold=True, size=9, color=WHITE)

    ws.merge_cells("F6:I6")
    apply(ws["F6"], "With Backside Covered (Test 2)", fill_color=BROWN_HDR, bold=True, size=9, color=WHITE)

    ws.merge_cells("J6:K6")
    apply(ws["J6"], "Delta (T1 - T2)", fill_color=PURPLE_HDR, bold=True, size=9, color=WHITE)
    ws.row_dimensions[6].height = 19.5

    # ── Row 7: Sub-headers ──
    for c, (v, fc) in enumerate([
        ("Measured", GREEN_SUB), ("Predicted", GREEN_SUB), ("STC", GREEN_SUB), ("Unit", GREEN_SUB),
        ("Measured", BROWN_SUB), ("Predicted", BROWN_SUB), ("STC", BROWN_SUB), ("Unit", BROWN_SUB),
        ("Meas Δ", PURPLE_SUB), ("STC Δ", PURPLE_SUB)
    ], 2):
        apply(ws.cell(7, c), v, fill_color=fc, bold=True, size=9, color=WHITE)
    ws.row_dimensions[7].height = 19.5

    def d(key_m, key_s=None):
        try:
            v1 = t1.get(key_m)
            v2 = t2.get(key_m)
            if v1 is not None and v2 is not None:
                return round(float(v1) - float(v2), 3)
            if key_s:
                v1s = t1.get(key_s)
                v2s = t2.get(key_s)
                if v1s is not None and v2s is not None:
                    return round(float(v1s) - float(v2s), 3)
        except (TypeError, ValueError):
            pass
        return "—"

    # Row data: (param, t1_meas, t1_pred, t1_stc, unit, t2_meas, t2_pred, t2_stc, unit, delta_meas, delta_stc, alt_rows)
    # alt_rows: True = alternating light blue, False = white
    rows_data = [
        ("Performance (%)",
         t1.get("performance","—"), "—", "—", "%",
         t2.get("performance","—"), "—", "—", "%",
         d("performance"), "—", True),

        ("Fill Factor",
         t1.get("fillFactor","—"), "0.74", "0.77", "",
         t2.get("fillFactor","—"), "0.74", "0.75", "",
         d("fillFactor"), "—", False),

        ("Pmax (W)",
         t1.get("pmaxMeasured","—"), t1.get("pmaxPredicted","—"), t1.get("pmaxSTC","—"), "W",
         t2.get("pmaxMeasured","—"), t2.get("pmaxPredicted","—"), t2.get("pmaxSTC","—"), "W",
         d("pmaxMeasured"), d("pmaxSTC"), True),

        ("Irr (W/m²)",
         t1.get("irradiance","—"), "—", 1000, "W/m²",
         t2.get("irradiance","—"), "—", 1000, "W/m²",
         d("irradiance"), "—", False),

        ("Isc (A)",
         t1.get("iscMeasured","—"), t1.get("iscPredicted","—"), t1.get("iscSTC","—"), "A",
         t2.get("iscMeasured","—"), t2.get("iscPredicted","—"), t2.get("iscSTC","—"), "A",
         d("iscMeasured"), d("iscSTC"), True),

        ("Cell Temp (°C)",
         t1.get("cellTemp","—"), "—", 25, "°C",
         t2.get("cellTemp","—"), "—", 25, "°C",
         d("cellTemp"), "—", False),

        ("Voc (V)",
         t1.get("vocMeasured","—"), t1.get("vocPredicted","—"), t1.get("vocSTC","—"), "V",
         t2.get("vocMeasured","—"), t2.get("vocPredicted","—"), t2.get("vocSTC","—"), "V",
         d("vocMeasured"), d("vocSTC"), True),

        ("Imp (A)",
         t1.get("impMeasured","—"), t1.get("impPredicted","—"), t1.get("impSTC","—"), "A",
         t2.get("impMeasured","—"), t2.get("impPredicted","—"), t2.get("impSTC","—"), "A",
         d("impMeasured"), d("impSTC"), False),

        ("Vmp (V)",
         t1.get("vmpMeasured","—"), t1.get("vmpPredicted","—"), t1.get("vmpSTC","—"), "V",
         t2.get("vmpMeasured","—"), t2.get("vmpPredicted","—"), t2.get("vmpSTC","—"), "V",
         d("vmpMeasured"), d("vmpSTC"), True),

        ("Current Ratio",
         t1.get("currentRatioMeas","—"), t1.get("currentRatioPred","—"), t1.get("currentRatioSTC","—"), "",
         t2.get("currentRatioMeas","—"), t2.get("currentRatioPred","—"), t2.get("currentRatioSTC","—"), "",
         d("currentRatioMeas"), "—", False),

        ("Voltage Ratio",
         t1.get("voltageRatioMeas","—"), t1.get("voltageRatioPred","—"), t1.get("voltageRatioSTC","—"), "",
         t2.get("voltageRatioMeas","—"), t2.get("voltageRatioPred","—"), t2.get("voltageRatioSTC","—"), "",
         d("voltageRatioMeas"), "—", True),
    ]

    for ri, rd in enumerate(rows_data):
        r = ri + 8
        param, t1m, t1p, t1s, u1, t2m, t2p, t2s, u2, dm, ds, alt = rd
        alt_fill = DATA_LIGHT if alt else WHITE

        # Column A: parameter label
        c = ws.cell(r, 1)
        apply(c, param, fill_color=PARAM_BLUE, bold=True, size=9, color=BLACK, align="left")

        # T1 data
        for ci, v in enumerate([t1m, t1p, t1s, u1], 2):
            apply(ws.cell(r, ci), v, fill_color=alt_fill, size=9, color=BLACK)

        # T2 data
        for ci, v in enumerate([t2m, t2p, t2s, u2], 6):
            apply(ws.cell(r, ci), v, fill_color=alt_fill, size=9, color=BLACK)

        # Delta: color positive red, negative green
        for ci, v in enumerate([dm, ds], 10):
            delta_color = BLACK
            if isinstance(v, (int, float)):
                delta_color = RED_FONT if v > 0 else DARK_GREEN_FONT if v < 0 else BLACK
            apply(ws.cell(r, ci), v, fill_color=DELTA_BG, bold=True, size=9, color=delta_color)

        ws.row_dimensions[r].height = 19.5

    # ── Assessment row ──
    assess_row = len(rows_data) + 8
    ws.merge_cells(f"A{assess_row}:K{assess_row+1}")
    assessment_text = f"📋 ASSESSMENT: {m.get('assessment', 'No assessment recorded.')}"
    c = ws.cell(assess_row, 1)
    c.value = assessment_text
    c.fill = fill(ASSESS_BG)
    c.font = Font(bold=False, size=9, color=BLUE_FONT, name="Arial")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border = thin_border()
    ws.row_dimensions[assess_row].height = 49.5
    ws.row_dimensions[assess_row+1].height = 30

    # ── Column widths ──
    col_widths = {"A":20,"B":12,"C":12,"D":12,"E":8,"F":12,"G":12,"H":12,"I":8,"J":12,"K":12}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "B8"


def generate(modules, output_path, test_date=None):
    from datetime import date
    if not test_date:
        test_date = date.today().strftime("%d/%m/%Y")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_summary(wb, modules, test_date)

    for i, m in enumerate(modules, 1):
        build_module_sheet(wb, m, i)

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 generate_report.py modules.json output.xlsx")
        sys.exit(1)
    with open(sys.argv[1]) as f:
        modules = json.load(f)
    out = generate(modules, sys.argv[2])
    print(f"Saved: {out}")
